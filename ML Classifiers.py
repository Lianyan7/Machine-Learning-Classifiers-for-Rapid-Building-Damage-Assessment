#!/usr/bin/env python3
"""
Integrated Classifier Evaluation Script
=========================================

This script reads data from an Excel workbook ("RBA.xlsx") containing four sheets,
performs data preprocessing, and evaluates four classifiers (KNN, MLP, Random Forest,
and SVM) over a range of hyperparameters. For each classifier, training and testing
results (including confusion matrices, F1 score, recall, precision, accuracy, and timing)
are saved to separate Excel workbooks in the "Detailed results" directory.

A confusion matrix visualisation (with additional indicators) is generated using the MLP
classifier's testing confusion matrix from the final parameter combination.

Author: [Your Name]
Date: [Date]
License: [Appropriate License]
"""

import os
import time
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.colors import ListedColormap
from openpyxl import Workbook

from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (accuracy_score, confusion_matrix, f1_score,
                             recall_score, precision_score)

# =============================================================================
# DATA PREPROCESSING
# =============================================================================

# File path to the Excel workbook containing the four sheets.
FILE_PATH = 'RBA.xlsx'

# Read data from the Excel file
df_value = pd.read_excel(FILE_PATH, sheet_name='Value')
df_multiclass = pd.read_excel(FILE_PATH, sheet_name='Multi-class')
df_binary = pd.read_excel(FILE_PATH, sheet_name='Binary')
df_label = pd.read_excel(FILE_PATH, sheet_name='Label')

# Assign features and labels
X_value = df_value
X_multiclass = df_multiclass
X_binary = df_binary
y = df_label

# Standardise continuous features
scaler = StandardScaler()
X_value_normalised = scaler.fit_transform(X_value)

# One-hot encode categorical features
X_multiclass_encoded = pd.get_dummies(X_multiclass)
X_binary_encoded = pd.get_dummies(X_binary)

# One-hot encode labels (for KNN, MLP, and Random Forest)
y_encoded = pd.get_dummies(y)

# Combine all features into a single feature matrix
X_combined = np.concatenate([X_value_normalised, X_multiclass_encoded, X_binary_encoded], axis=1)

# Split the data into training (80%) and testing (20%) sets
X_train, X_test, y_train_encoded, y_test_encoded = train_test_split(
    X_combined, y_encoded, test_size=0.2, random_state=42
)

# For classifiers (e.g. SVM) that require 1D labels, extract class indices
y_train_class = y_train_encoded.values.argmax(axis=1)
y_test_class = y_test_encoded.values.argmax(axis=1)

# Create output directory if it does not exist
OUTPUT_DIR = 'Detailed results'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# =============================================================================
# 1. K-NEAREST NEIGHBOURS (KNN) CLASSIFIER
# =============================================================================

neighbor_values = range(1, 11)
results_file_path_knn = os.path.join(OUTPUT_DIR, 'KNN_results.xlsx')
wb_knn = Workbook()

# Create sheets for training and testing results
ws_train_knn = wb_knn.active
ws_train_knn.title = "Training Results"
ws_train_knn.append(['Neighbors', 'Confusion Matrix', 'F1 Score', 'Recall',
                     'Precision', 'Accuracy', 'Training Time (s)', 'Testing Time (s)'])

ws_test_knn = wb_knn.create_sheet(title="Testing Results")
ws_test_knn.append(['Neighbors', 'Confusion Matrix', 'F1 Score', 'Recall',
                    'Precision', 'Accuracy'])

for n_neighbors in neighbor_values:
    knn = KNeighborsClassifier(n_neighbors=n_neighbors)

    # Train KNN classifier and record training time
    start_time = time.time()
    knn.fit(X_train, y_train_encoded)
    training_time = time.time() - start_time

    # Predict on the testing set and record testing time
    start_time = time.time()
    y_test_pred = knn.predict(X_test)
    testing_time = time.time() - start_time

    # Evaluate on training set (convert one-hot predictions to class indices)
    y_train_pred = knn.predict(X_train)
    y_train_true = y_train_encoded.values.argmax(axis=1)
    y_train_pred_class = y_train_pred.argmax(axis=1)

    train_conf_matrix = confusion_matrix(y_train_true, y_train_pred_class)
    train_f1 = f1_score(y_train_true, y_train_pred_class, average='weighted')
    train_recall = recall_score(y_train_true, y_train_pred_class, average='weighted')
    train_precision = precision_score(y_train_true, y_train_pred_class, average='weighted')
    train_accuracy = accuracy_score(y_train_true, y_train_pred_class)

    ws_train_knn.append([n_neighbors, str(train_conf_matrix), train_f1,
                         train_recall, train_precision, train_accuracy,
                         training_time, testing_time])

    # Evaluate on testing set
    y_test_true = y_test_encoded.values.argmax(axis=1)
    y_test_pred_class = y_test_pred.argmax(axis=1)
    test_conf_matrix = confusion_matrix(y_test_true, y_test_pred_class)
    test_f1 = f1_score(y_test_true, y_test_pred_class, average='weighted')
    test_recall = recall_score(y_test_true, y_test_pred_class, average='weighted')
    test_precision = precision_score(y_test_true, y_test_pred_class, average='weighted')
    test_accuracy = accuracy_score(y_test_true, y_test_pred_class)

    ws_test_knn.append([n_neighbors, str(test_conf_matrix), test_f1,
                        test_recall, test_precision, test_accuracy])

    print(f"KNN (Neighbors = {n_neighbors}) - Test Confusion Matrix:\n{test_conf_matrix}")

wb_knn.save(results_file_path_knn)
print(f"KNN results saved to {results_file_path_knn}")

# =============================================================================
# 2. MULTI-LAYER PERCEPTRON (MLP) CLASSIFIER
# =============================================================================

hidden_states = [100, 200, 300, 400, 500]
max_iterations = [100, 300, 500, 700]
learning_rates = [0.01, 0.001]
results_file_path_mlp = os.path.join(OUTPUT_DIR, 'MLP_results_summary.xlsx')
wb_mlp = Workbook()

ws_train_mlp = wb_mlp.active
ws_train_mlp.title = "Training Results"
ws_train_mlp.append(['Hidden Layer Size', 'Max Iterations', 'Learning Rate',
                     'Confusion Matrix', 'F1 Score', 'Recall',
                     'Precision', 'Accuracy', 'Training Time (s)'])

ws_test_mlp = wb_mlp.create_sheet(title="Testing Results")
ws_test_mlp.append(['Hidden Layer Size', 'Max Iterations', 'Learning Rate',
                    'Confusion Matrix', 'F1 Score', 'Recall',
                    'Precision', 'Accuracy', 'Testing Time (s)'])

# Store one confusion matrix for visualisation (from final MLP iteration)
mlp_confusion_matrix_for_plot = None

for hidden_state in hidden_states:
    for max_iter in max_iterations:
        for learning_rate in learning_rates:
            mlp_classifier = MLPClassifier(hidden_layer_sizes=(hidden_state,),
                                           max_iter=max_iter,
                                           learning_rate_init=learning_rate,
                                           random_state=42)
            start_time = time.time()
            mlp_classifier.fit(X_train, y_train_encoded)
            training_time = time.time() - start_time

            start_time = time.time()
            y_test_pred = mlp_classifier.predict(X_test)
            testing_time = time.time() - start_time

            y_train_pred = mlp_classifier.predict(X_train)
            train_conf_matrix = confusion_matrix(y_train_encoded.values.argmax(axis=1),
                                                 y_train_pred.argmax(axis=1))
            train_f1 = f1_score(y_train_encoded, y_train_pred, average='weighted')
            train_recall = recall_score(y_train_encoded, y_train_pred, average='weighted')
            train_precision = precision_score(y_train_encoded, y_train_pred, average='weighted')
            train_accuracy = accuracy_score(y_train_encoded, y_train_pred)

            ws_train_mlp.append([hidden_state, max_iter, learning_rate, str(train_conf_matrix),
                                 train_f1, train_recall, train_precision, train_accuracy, training_time])

            test_conf_matrix = confusion_matrix(y_test_encoded.values.argmax(axis=1),
                                                y_test_pred.argmax(axis=1))
            test_f1 = f1_score(y_test_encoded, y_test_pred, average='weighted')
            test_recall = recall_score(y_test_encoded, y_test_pred, average='weighted')
            test_precision = precision_score(y_test_encoded, y_test_pred, average='weighted')
            test_accuracy = accuracy_score(y_test_encoded, y_test_pred)

            ws_test_mlp.append([hidden_state, max_iter, learning_rate, str(test_conf_matrix),
                                test_f1, test_recall, test_precision, test_accuracy, testing_time])

            print(
                f"MLP (Hidden: {hidden_state}, Iterations: {max_iter}, LR: {learning_rate}) - Test Confusion Matrix:\n{test_conf_matrix}")

            # Update the confusion matrix for plotting
            mlp_confusion_matrix_for_plot = test_conf_matrix

wb_mlp.save(results_file_path_mlp)
print(f"MLP results saved to {results_file_path_mlp}")

# =============================================================================
# 3. RANDOM FOREST (RF) CLASSIFIER
# =============================================================================

n_estimators_values = [50, 60, 70, 80, 90, 100, 110, 120]
results_file_path_rf = os.path.join(OUTPUT_DIR, 'RandomForest_results_summary.xlsx')
wb_rf = Workbook()

ws_train_rf = wb_rf.active
ws_train_rf.title = "Training Results"
ws_train_rf.append(['n_estimators', 'Confusion Matrix', 'F1 Score',
                    'Recall', 'Precision', 'Accuracy', 'Training Time (s)'])

ws_test_rf = wb_rf.create_sheet(title="Testing Results")
ws_test_rf.append(['n_estimators', 'Confusion Matrix', 'F1 Score',
                   'Recall', 'Precision', 'Accuracy', 'Testing Time (s)'])

for n_estimators in n_estimators_values:
    rf_classifier = RandomForestClassifier(n_estimators=n_estimators, random_state=42)
    start_time = time.time()
    rf_classifier.fit(X_train, y_train_encoded)
    training_time = time.time() - start_time

    start_time = time.time()
    y_test_pred = rf_classifier.predict(X_test)
    testing_time = time.time() - start_time

    y_train_pred = rf_classifier.predict(X_train)
    train_conf_matrix = confusion_matrix(y_train_encoded.values.argmax(axis=1),
                                         y_train_pred.argmax(axis=1))
    train_f1 = f1_score(y_train_encoded, y_train_pred, average='weighted')
    train_recall = recall_score(y_train_encoded, y_train_pred, average='weighted')
    train_precision = precision_score(y_train_encoded, y_train_pred, average='weighted')
    train_accuracy = accuracy_score(y_train_encoded, y_train_pred)

    ws_train_rf.append([n_estimators, str(train_conf_matrix), train_f1,
                        train_recall, train_precision, train_accuracy, training_time])

    test_conf_matrix = confusion_matrix(y_test_encoded.values.argmax(axis=1),
                                        y_test_pred.argmax(axis=1))
    test_f1 = f1_score(y_test_encoded, y_test_pred, average='weighted')
    test_recall = recall_score(y_test_encoded, y_test_pred, average='weighted')
    test_precision = precision_score(y_test_encoded, y_test_pred, average='weighted')
    test_accuracy = accuracy_score(y_test_encoded, y_test_pred)

    ws_test_rf.append([n_estimators, str(test_conf_matrix), test_f1,
                       test_recall, test_precision, test_accuracy, testing_time])

    print(f"RF (n_estimators = {n_estimators}) - Test Confusion Matrix:\n{test_conf_matrix}")

wb_rf.save(results_file_path_rf)
print(f"Random Forest results saved to {results_file_path_rf}")

# =============================================================================
# 4. SUPPORT VECTOR MACHINE (SVM) CLASSIFIER
# =============================================================================

kernel_values = ['rbf', 'linear', 'poly']
degree_values = [2, 3, 4, 5]
gamma_values = ['scale', 'auto']
results_file_path_svm = os.path.join(OUTPUT_DIR, 'SVM_results_summary.xlsx')
wb_svm = Workbook()

ws_train_svm = wb_svm.active
ws_train_svm.title = "Training Results"
ws_train_svm.append(['Kernel', 'Degree', 'Gamma', 'Confusion Matrix',
                     'F1 Score', 'Recall', 'Precision', 'Accuracy', 'Training Time (s)'])

ws_test_svm = wb_svm.create_sheet(title="Testing Results")
ws_test_svm.append(['Kernel', 'Degree', 'Gamma', 'Confusion Matrix',
                    'F1 Score', 'Recall', 'Precision', 'Accuracy', 'Testing Time (s)'])

for kernel in kernel_values:
    for degree in degree_values:
        for gamma in gamma_values:
            svm_classifier = SVC(kernel=kernel, degree=degree, gamma=gamma, random_state=42)
            start_time = time.time()
            svm_classifier.fit(X_train, y_train_class)
            training_time = time.time() - start_time

            start_time = time.time()
            y_test_pred = svm_classifier.predict(X_test)
            testing_time = time.time() - start_time

            y_train_pred = svm_classifier.predict(X_train)
            train_conf_matrix = confusion_matrix(y_train_class, y_train_pred)
            train_f1 = f1_score(y_train_class, y_train_pred, average='weighted')
            train_recall = recall_score(y_train_class, y_train_pred, average='weighted')
            train_precision = precision_score(y_train_class, y_train_pred, average='weighted')
            train_accuracy = accuracy_score(y_train_class, y_train_pred)

            ws_train_svm.append([kernel, degree, gamma, str(train_conf_matrix), train_f1,
                                 train_recall, train_precision, train_accuracy, training_time])

            test_conf_matrix = confusion_matrix(y_test_class, y_test_pred)
            test_f1 = f1_score(y_test_class, y_test_pred, average='weighted')
            test_recall = recall_score(y_test_class, y_test_pred, average='weighted')
            test_precision = precision_score(y_test_class, y_test_pred, average='weighted')
            test_accuracy = accuracy_score(y_test_class, y_test_pred)

            ws_test_svm.append([kernel, degree, gamma, str(test_conf_matrix), test_f1,
                                test_recall, test_precision, test_accuracy, testing_time])

            print(
                f"SVM (Kernel: {kernel}, Degree: {degree}, Gamma: {gamma}) - Test Confusion Matrix:\n{test_conf_matrix}")

wb_svm.save(results_file_path_svm)
print(f"SVM results saved to {results_file_path_svm}")

# =============================================================================
# 5. CONFUSION MATRIX VISUALISATION
# =============================================================================

if mlp_confusion_matrix_for_plot is None:
    raise ValueError("No confusion matrix available from MLP testing for visualisation.")

# Define class labels (assumed 10 damage states)
states = ['G', 'Y', 'R', 'G1', 'G2', 'Y1', 'Y2', 'R1', 'R2', 'R3']

# Configure font properties for a professional appearance
plt.rcParams['font.family'] = 'Times New Roman'
plt.rcParams['font.size'] = 24

# Compute annotations for the heatmap
c_mat = mlp_confusion_matrix_for_plot
total = np.sum(c_mat)
labels = [[f"{val:0.0f}\n{val / total:.2%}" for val in row] for row in c_mat]

plt.figure(figsize=(15, 13))
plt.subplots_adjust(left=0.03, right=0.97, top=0.95, bottom=0.03)
ax = sns.heatmap(c_mat, annot=labels, cmap="OrRd", fmt='',
                 xticklabels=states, yticklabels=states, cbar=True,
                 annot_kws={"size": 24}, linewidths=0.5, linecolor='gray')

ax.set_title('Predicted Damage State', fontweight='bold', fontsize=22)
ax.tick_params(labeltop=True, labelbottom=False, length=0)
ax.set_ylabel('Actual Damage State', fontweight='bold', fontsize=18)
ax.set_xlabel('', fontsize=20)
ax.tick_params(axis='y', pad=10)

# Extra matrix to display recall, precision, and overall accuracy
f_mat = np.zeros((c_mat.shape[0] + 1, c_mat.shape[1] + 1))
with np.errstate(divide='ignore', invalid='ignore'):
    f_mat[:-1, -1] = np.diag(c_mat) / np.sum(c_mat, axis=1)  # Recall per class
    f_mat[-1, :-1] = np.diag(c_mat) / np.sum(c_mat, axis=0)  # Precision per class
    f_mat[-1, -1] = np.trace(c_mat) / total  # Overall accuracy
    f_mat = np.nan_to_num(f_mat)

f_mask = np.ones_like(f_mat)
f_mask[:, -1] = 0
f_mask[-1, :] = 0

annot_font_size = 24 if np.any(f_mat[:-1, -1] > 0.2) else 19
f_color = np.ones_like(f_mat)
f_color[-1, -1] = 0
f_annot = [[f"{val:0.2%}" for val in row] for row in f_mat]
f_annot[-1][-1] = "Acc.:\n" + f_annot[-1][-1]

sns.heatmap(f_color, mask=f_mask, annot=f_annot, fmt='',
            xticklabels=states + ["Recall"],
            yticklabels=states + ["Precision"],
            cmap=ListedColormap(['#f7f7f7', '#d9d9d9']),
            cbar=False, ax=ax,
            annot_kws={"size": annot_font_size})

yticklabels = ax.get_yticklabels()
xticklabels = ax.get_xticklabels()
yticklabels[-1].set_size(24)
xticklabels[-1].set_size(24)
ax.set_yticklabels(yticklabels, rotation=90, ha='center')
ax.set_xticklabels(xticklabels, rotation=0, ha='center')

plt.show()
