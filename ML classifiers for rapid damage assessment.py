"""
Integrated Classifier Evaluation Script (6 classifiers)
======================================================

Classifiers:
  1) KNN
  2) MLP
  3) Random Forest (RF)
  4) SVM
  5) Gradient Boosting Classifier (GBC)
  6) Gradient Bagging (GBag) 

Evaluation design (engineering-credible and leakage-safe):
  - Outer split: stratified 80/20 hold-out test set.
  - Inner validation: stratified 5-fold CV within the 80% training set via GridSearchCV.

Outputs:
  - One Excel workbook per classifier in "Detailed results":
      * "CV Summary" sheet: best params, best CV score, protocol metadata
      * "Test Results" sheet: hold-out confusion matrix + metrics + timings
      * Optional "CV Grid" sheet: full GridSearchCV table
  - Confusion-matrix visualisation for the best MLP (selected by CV f1_weighted),
    evaluated on the held-out test set, with added Recall/Precision/Accuracy indicators.(just given an example)

Author: [Lianyan Li]
Date: [22/02/2025]
"""

import os
import time
import argparse
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.colors import ListedColormap
from openpyxl import Workbook

from sklearn.model_selection import train_test_split, StratifiedKFold, GridSearchCV
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import StandardScaler, OneHotEncoder, FunctionTransformer
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score, recall_score, precision_score

from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, BaggingClassifier
from sklearn.svm import SVC


# =============================================================================
# CONFIG
# =============================================================================

RANDOM_STATE = 42
OUTPUT_DIR_DEFAULT = "Detailed results"
PREFERRED_STATES = ["G", "Y", "R", "G1", "G2", "Y1", "Y2", "R1", "R2", "R3"]


# =============================================================================
# HELPERS
# =============================================================================

def _dense_if_sparse(X):
    """Convert sparse matrices to dense arrays; pass-through for dense inputs."""
    return X.toarray() if hasattr(X, "toarray") else X


def parse_int_list(s):
    return [int(x.strip()) for x in s.split(",") if x.strip()]


def parse_float_list(s):
    return [float(x.strip()) for x in s.split(",") if x.strip()]


def build_preprocess(numeric_cols, categorical_cols):
    """Leakage-safe preprocessing."""
    # Compatibility across scikit-learn versions
    try:
        ohe = OneHotEncoder(handle_unknown="ignore", sparse_output=True)
    except TypeError:
        ohe = OneHotEncoder(handle_unknown="ignore", sparse=True)

    return ColumnTransformer(
        transformers=[
            ("num", StandardScaler(), numeric_cols),
            ("cat", ohe, categorical_cols),
        ],
        remainder="drop",
    )


def load_raw_data(file_path):
    """
    Load four sheets and return:
      - X_df: combined raw DataFrame
      - y: 1D array of labels (string)
      - numeric_cols, categorical_cols
    """
    df_value = pd.read_excel(file_path, sheet_name="Value")
    df_multi = pd.read_excel(file_path, sheet_name="Multi-class")
    df_binary = pd.read_excel(file_path, sheet_name="Binary")
    df_label = pd.read_excel(file_path, sheet_name="Label")

    if df_label.shape[1] < 1:
        raise ValueError("Sheet 'Label' must contain at least one column.")
    if df_label.shape[1] > 1:
        print("Warning: 'Label' sheet has multiple columns; using the first column as y.")

    X_df = pd.concat([df_value, df_multi, df_binary], axis=1)
    y = df_label.iloc[:, 0].astype(str).values

    numeric_cols = df_value.columns.tolist()
    categorical_cols = df_multi.columns.tolist() + df_binary.columns.tolist()

    return X_df, y, numeric_cols, categorical_cols


def get_label_order(y_all):
    """Use preferred engineering label order if possible; else use sorted unique labels."""
    labels_unique = sorted(pd.unique(pd.Series(y_all).astype(str)))
    if set(PREFERRED_STATES).issubset(set(labels_unique)):
        extras = [lab for lab in labels_unique if lab not in PREFERRED_STATES]
        return PREFERRED_STATES + extras
    return labels_unique


def compute_metrics(y_true, y_pred, labels_order):
    """Compute confusion matrix + weighted metrics."""
    cm = confusion_matrix(y_true, y_pred, labels=labels_order)
    return {
        "confusion_matrix": cm,
        "f1_weighted": float(f1_score(y_true, y_pred, average="weighted", zero_division=0)),
        "recall_weighted": float(recall_score(y_true, y_pred, average="weighted", zero_division=0)),
        "precision_weighted": float(precision_score(y_true, y_pred, average="weighted", zero_division=0)),
        "accuracy": float(accuracy_score(y_true, y_pred)),
    }


def evaluate_on_holdout(best_estimator, X_train, y_train, X_test, y_test, labels_order):
    """Fit on full train80 and evaluate once on test20; return metrics and timings."""
    t0 = time.time()
    best_estimator.fit(X_train, y_train)
    fit_time = time.time() - t0

    t0 = time.time()
    y_test_pred = best_estimator.predict(X_test)
    pred_time = time.time() - t0

    test_metrics = compute_metrics(y_test, y_test_pred, labels_order)

    # Optional: training metrics (diagnostic)
    y_train_pred = best_estimator.predict(X_train)
    train_metrics = compute_metrics(y_train, y_train_pred, labels_order)

    return {
        **test_metrics,
        "train_f1": train_metrics["f1_weighted"],
        "train_precision": train_metrics["precision_weighted"],
        "train_recall": train_metrics["recall_weighted"],
        "train_accuracy": train_metrics["accuracy"],
        "fit_time_s": float(fit_time),
        "pred_time_s": float(pred_time),
    }


def save_excel_results(filepath, cv_info, test_results, cv_grid_df=None):
    """Save CV summary and hold-out test results to Excel. Optionally include full CV grid."""
    wb = Workbook()
    ws_cv = wb.active
    ws_cv.title = "CV Summary"
    ws_cv.append(["Item", "Value"])
    for k, v in cv_info.items():
        ws_cv.append([k, str(v)])

    ws_test = wb.create_sheet(title="Test Results")
    ws_test.append(["Metric", "Value"])
    for k, v in test_results.items():
        ws_test.append([k, str(v)])

    if cv_grid_df is not None:
        ws_grid = wb.create_sheet(title="CV Grid")
        ws_grid.append(list(cv_grid_df.columns))
        for row in cv_grid_df.itertuples(index=False):
            ws_grid.append(list(row))

    wb.save(filepath)
    print(f"Results saved to: {filepath}")


def gridsearch_fit(pipe, param_grid, X_train, y_train, inner_cv, n_jobs, scoring="f1_weighted"):
    """Run GridSearchCV on train80 only."""
    gs = GridSearchCV(
        estimator=pipe,
        param_grid=param_grid,
        scoring=scoring,
        cv=inner_cv,
        n_jobs=n_jobs,
        refit=True,
        return_train_score=False,
    )
    gs.fit(X_train, y_train)
    return gs


def plot_confusion_matrix_with_indicators(c_mat, labels_order, title):
    """Plot confusion matrix + Recall/Precision/Accuracy indicators (test set)."""
    total = np.sum(c_mat)

    plt.rcParams["font.family"] = "Times New Roman"
    plt.rcParams["font.size"] = 24

    ann = [[f"{val:0.0f}\n{(val / total):.2%}" if total > 0 else f"{val:0.0f}\n0.00%"
            for val in row] for row in c_mat]

    plt.figure(figsize=(15, 13))
    plt.subplots_adjust(left=0.03, right=0.97, top=0.95, bottom=0.03)

    ax = sns.heatmap(
        c_mat, annot=ann, cmap="OrRd", fmt="",
        xticklabels=labels_order, yticklabels=labels_order, cbar=True,
        annot_kws={"size": 24}, linewidths=0.5, linecolor="gray"
    )

    ax.set_title(title, fontweight="bold", fontsize=20)
    ax.tick_params(labeltop=True, labelbottom=False, length=0)
    ax.set_ylabel("Actual Class", fontweight="bold", fontsize=18)
    ax.set_xlabel("")
    ax.tick_params(axis="y", pad=10)

    f_mat = np.zeros((c_mat.shape[0] + 1, c_mat.shape[1] + 1), dtype=float)
    with np.errstate(divide="ignore", invalid="ignore"):
        row_sums = np.sum(c_mat, axis=1)
        col_sums = np.sum(c_mat, axis=0)
        diag = np.diag(c_mat)

        f_mat[:-1, -1] = np.divide(diag, row_sums, out=np.zeros_like(diag, dtype=float), where=row_sums != 0)
        f_mat[-1, :-1] = np.divide(diag, col_sums, out=np.zeros_like(diag, dtype=float), where=col_sums != 0)
        f_mat[-1, -1] = (np.trace(c_mat) / total) if total > 0 else 0.0

    f_mask = np.ones_like(f_mat, dtype=bool)
    f_mask[:, -1] = False
    f_mask[-1, :] = False

    f_color = np.ones_like(f_mat)
    f_color[-1, -1] = 0

    f_annot = [[f"{val:0.2%}" for val in row] for row in f_mat]
    f_annot[-1][-1] = "Acc.:\n" + f_annot[-1][-1]

    sns.heatmap(
        f_color, mask=f_mask, annot=f_annot, fmt="",
        xticklabels=labels_order + ["Recall"],
        yticklabels=labels_order + ["Precision"],
        cmap=ListedColormap(["#f7f7f7", "#d9d9d9"]),
        cbar=False, ax=ax,
        annot_kws={"size": 22}
    )

    yticklabels = ax.get_yticklabels()
    xticklabels = ax.get_xticklabels()
    yticklabels[-1].set_size(24)
    xticklabels[-1].set_size(24)
    ax.set_yticklabels(yticklabels, rotation=90, ha="center")
    ax.set_xticklabels(xticklabels, rotation=0, ha="center")

    plt.show()


# =============================================================================
# MAIN
# =============================================================================

def main(args):
    os.makedirs(args.output_dir, exist_ok=True)

    X_df, y, numeric_cols, categorical_cols = load_raw_data(args.input)
    labels_order = get_label_order(y)

    # Outer split: stratified 80/20 hold-out
    X_train, X_test, y_train, y_test = train_test_split(
        X_df, y,
        test_size=args.test_size,
        random_state=RANDOM_STATE,
        stratify=y
    )

    # Inner CV: stratified K-fold on train80 only
    inner_cv = StratifiedKFold(
        n_splits=args.cv_folds,
        shuffle=True,
        random_state=RANDOM_STATE
    )

    preprocess = build_preprocess(numeric_cols, categorical_cols)
    to_dense = FunctionTransformer(_dense_if_sparse, accept_sparse=True)

    timestamp = time.strftime("%Y%m%d_%H%M%S")

    models = {}

    # 1) KNN
    knn_pipe = Pipeline([("prep", preprocess), ("dense", to_dense), ("clf", KNeighborsClassifier())])
    models["KNN"] = {
        "pipe": knn_pipe,
        "param_grid": {"clf__n_neighbors": list(range(1, 11))}
    }

    # 2) MLP
    mlp_pipe = Pipeline([("prep", preprocess), ("dense", to_dense),
                         ("clf", MLPClassifier(random_state=RANDOM_STATE))])
    models["MLP"] = {
        "pipe": mlp_pipe,
        "param_grid": {
            "clf__hidden_layer_sizes": [(100,), (200,), (300,), (400,), (500,)],
            "clf__max_iter": [100, 300, 500, 700],
            "clf__learning_rate_init": [0.01, 0.001],
        }
    }

    # 3) Random Forest
    rf_pipe = Pipeline([("prep", preprocess), ("dense", to_dense),
                        ("clf", RandomForestClassifier(random_state=RANDOM_STATE))])
    models["RF"] = {
        "pipe": rf_pipe,
        "param_grid": {"clf__n_estimators": [50, 60, 70, 80, 90, 100, 110, 120]}
    }

    # 4) SVM
    svm_pipe = Pipeline([("prep", preprocess), ("dense", to_dense), ("clf", SVC(random_state=RANDOM_STATE))])
    models["SVM"] = {
        "pipe": svm_pipe,
        "param_grid": {
            "clf__kernel": ["rbf", "linear", "poly"],
            "clf__degree": [2, 3, 4, 5],
            "clf__gamma": ["scale", "auto"],
        }
    }

    # 5) Gradient Boosting Classifier (GBC)
    gbc_pipe = Pipeline([("prep", preprocess), ("dense", to_dense),
                         ("clf", GradientBoostingClassifier(random_state=RANDOM_STATE))])
    models["GBC"] = {
        "pipe": gbc_pipe,
        "param_grid": {
            "clf__n_estimators": parse_int_list(args.gbc_n_estimators),
            "clf__learning_rate": parse_float_list(args.gbc_learning_rate),
            "clf__max_depth": parse_int_list(args.gbc_max_depth),
        }
    }

    # 6) Gradient Bagging (GBag)
    base_gbc = GradientBoostingClassifier(random_state=RANDOM_STATE)
    gbag_clf = BaggingClassifier(estimator=base_gbc, random_state=RANDOM_STATE)
    gbag_pipe = Pipeline([("prep", preprocess), ("dense", to_dense), ("clf", gbag_clf)])
    models["GBag"] = {
        "pipe": gbag_pipe,
        "param_grid": {
            "clf__estimator__n_estimators": parse_int_list(args.gbc_n_estimators),
            "clf__estimator__learning_rate": parse_float_list(args.gbc_learning_rate),
            "clf__estimator__max_depth": parse_int_list(args.gbc_max_depth),
            "clf__n_estimators": parse_int_list(args.gbag_n_estimators),
        }
    }

    # Track best MLP by CV score for plotting
    best_mlp_cv_score = -1.0
    best_mlp_params = None
    best_mlp_test_cm = None

    for model_name, cfg in models.items():
        print(f"\n=== Running {model_name} (CV inside train80, then hold-out test) ===")

        gs = gridsearch_fit(
            pipe=cfg["pipe"],
            param_grid=cfg["param_grid"],
            X_train=X_train,
            y_train=y_train,
            inner_cv=inner_cv,
            n_jobs=args.n_jobs,
            scoring="f1_weighted"
        )

        best_est = gs.best_estimator_
        cv_best_score = float(gs.best_score_)
        best_params = gs.best_params_

        print(f"{model_name} best params: {best_params}")
        print(f"{model_name} best CV f1_weighted: {cv_best_score:.6f}")

        test_results = evaluate_on_holdout(best_est, X_train, y_train, X_test, y_test, labels_order=labels_order)

        # Store best MLP for plotting
        if model_name == "MLP" and cv_best_score > best_mlp_cv_score:
            best_mlp_cv_score = cv_best_score
            best_mlp_params = best_params
            best_mlp_test_cm = test_results["confusion_matrix"]

        cv_grid_df = None
        if args.save_cv_grid:
            cv_grid_df = pd.DataFrame(gs.cv_results_)
            keep_cols = [c for c in cv_grid_df.columns if c.startswith("param_")] + [
                "mean_test_score", "std_test_score", "rank_test_score"
            ]
            keep_cols = [c for c in keep_cols if c in cv_grid_df.columns]
            cv_grid_df = cv_grid_df[keep_cols].sort_values("rank_test_score")

        cv_info = {
            "model": model_name,
            "best_params": best_params,
            "best_cv_f1_weighted": cv_best_score,
            "cv_folds": args.cv_folds,
            "outer_split": f"{int((1 - args.test_size) * 100)}/{int(args.test_size * 100)} stratified hold-out",
            "scoring": "f1_weighted",
            "random_state": RANDOM_STATE,
        }

        outfile = os.path.join(args.output_dir, f"{model_name}_results_holdout80_CV{args.cv_folds}_{timestamp}.xlsx")
        save_excel_results(outfile, cv_info, test_results, cv_grid_df=cv_grid_df)

    # Plot best MLP confusion matrix on hold-out test set
    if best_mlp_test_cm is not None:
        title = f"MLP hold-out test confusion matrix (best CV f1_weighted = {best_mlp_cv_score:.4f})"
        if best_mlp_params is not None:
            title += f"\nBest params: {best_mlp_params}"
        plot_confusion_matrix_with_indicators(best_mlp_test_cm, labels_order, title=title)
    else:
        print("No MLP result available for confusion matrix visualisation.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Evaluate 6 classifiers on RBA data (stratified 80/20 hold-out + stratified K-fold CV inside train80)."
    )
    parser.add_argument("--input", type=str, default="RBA.xlsx",
                        help="Path to Excel file with sheets: Value, Multi-class, Binary, Label")
    parser.add_argument("--output_dir", type=str, default=OUTPUT_DIR_DEFAULT,
                        help="Output directory for Excel results")
    parser.add_argument("--test_size", type=float, default=0.2,
                        help="Hold-out test proportion (default 0.2 for 80/20)")
    parser.add_argument("--cv_folds", type=int, default=5,
                        help="Number of inner CV folds (default 5)")
    parser.add_argument("--n_jobs", type=int, default=-1,
                        help="Parallel jobs for GridSearchCV (-1 uses all cores)")
    parser.add_argument("--save_cv_grid", action="store_true",
                        help="If set, saves full CV grid results to an extra sheet (can be large)")

    # GBC grids (accept both --gbc_* and legacy --gb_*)
    parser.add_argument("--gbc_n_estimators", "--gb_n_estimators", dest="gbc_n_estimators",
                        type=str, default="50,100,150",
                        help="Comma-separated GBC n_estimators grid")
    parser.add_argument("--gbc_learning_rate", "--gb_learning_rate", dest="gbc_learning_rate",
                        type=str, default="0.01,0.1,0.2",
                        help="Comma-separated GBC learning_rate grid")
    parser.add_argument("--gbc_max_depth", "--gb_max_depth", dest="gbc_max_depth",
                        type=str, default="3,4,5",
                        help="Comma-separated GBC max_depth grid")

    # Bagging grid
    parser.add_argument("--gbag_n_estimators", type=str, default="10,20,30",
                        help="Comma-separated Bagging n_estimators grid (outer bagging)")

    args = parser.parse_args()
    main(args)
